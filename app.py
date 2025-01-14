from flask import Flask, request, jsonify, render_template, make_response, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import os
import datetime
import logging
from collections import deque
import re

# Directories
RESULTS_DIRECTORY = os.path.join('data', 'resultados.csv')
LOG_DIRECTORY = os.path.join('data', 'logs.csv')
INFO_LOGGER = os.path.join(os.path.join('data', 'system_logs'), 'info.log')
DEBUG_LOGGER = os.path.join(os.path.join('data', 'system_logs'), 'debug.log')
RESULTS_LIST_HISTORY = os.path.join('data', 'results')
BACKUP_DIRECTORY = os.path.join('data', '')

history_stack = deque(maxlen=10)

# Loggers
logger_info = logging.getLogger('info_logger')
logger_info.setLevel(logging.INFO)

logger_debug = logging.getLogger('debug_logger')
logger_debug.setLevel(logging.DEBUG)

info_handler = logging.FileHandler(INFO_LOGGER, encoding='utf-8', errors='replace')
debug_handler = logging.FileHandler(DEBUG_LOGGER, encoding='utf-8', errors='replace')

info_handler.setLevel(logging.INFO)
debug_handler.setLevel(logging.DEBUG)

info_format = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
debug_format = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

info_handler.setFormatter(info_format)
debug_handler.setFormatter(debug_format)

logger_info.addHandler(info_handler)
logger_debug.addHandler(debug_handler)


app = Flask(__name__, static_folder='static', template_folder='templates')


@app.route('/')
def home():
    return render_template('index.html')


def load_df(directory):
    try:
        if os.path.exists(directory):
            df = pd.read_csv(directory)
            logger_info.info(f"DataFrame cargado correctamente desde {directory}")
        else:
            df = pd.DataFrame(columns=['Jugador', 'Puntos'])
            logger_info.info(f"El archivo {directory} no existe. Se creó un DataFrame vacío.")
        return df
    except Exception as e:
        logger_debug.error(f"Error loading DataFrame from {directory}: {e}")
        return pd.DataFrame(columns=['Jugador', 'Puntos'])
    

def save_df(df, ruta):
    try:
        df.to_csv(ruta, index=False)
        logger_info.info(f"DataFrame guardado correctamente en {ruta}")
    except Exception as e:
        logger_debug.error(f"Error saving DataFrame to {ruta}: {e}")


def delete_points(jugador):
    try:
        if os.path.exists(RESULTS_DIRECTORY):
            df = pd.read_csv(RESULTS_DIRECTORY)
            df.loc[df['Jugador'] == jugador, 'Puntos'] = 0
            df.loc[df['Jugador'] == jugador, 'Victorias'] += 1
            df.to_csv(RESULTS_DIRECTORY, index=False)
            logger_info.info(f"Puntos eliminados para el jugador {jugador}")
        else:
            logger_debug.warning("El archivo de resultados no existe.")
    except Exception as e:
        logger_debug.error(f"Error al eliminar puntos para {jugador}: {e}")


def calcular_puntos(jugadas, jugador):
    puntos = []
    tipos = []
    for jugada in jugadas:
        jugada = jugada.lower()

        if 'jugador' in jugada.lower() or 'jugadora' in jugada.lower():
            # Si la jugada contiene 'jugador', no se suman puntos y se ignora.
            tipos.append('Jugador')
            puntos.append(0)  # Puntos para jugada ignorada
            delete_points(jugador)
            continue
        elif 'nuevo' in jugada.lower():
            puntos.append(5)
            tipos.append('Nuevo')
        elif 'intercambio' in jugada.lower():
            puntos.append(3)
            tipos.append('Intercambio')
        else:
            puntos.append(1)
            tipos.append('Regular')

    logger_info.info(f"Puntos calculados para el jugador {jugador}: {puntos}")
    logger_info.info(f"Tipos calculados para el jugador {jugador}: {tipos}")
    return puntos, tipos


def create_log(types):
    try:
        # Verificar si el archivo existe
        if os.path.exists(LOG_DIRECTORY):
            try:
                log_df = pd.read_csv(LOG_DIRECTORY)
                if log_df.empty:
                    log_df = pd.DataFrame(columns=['Jugador', 'Fecha', 'Hora', 'Puntos', 'Tipo'])
            except pd.errors.EmptyDataError:
                log_df = pd.DataFrame(columns=['Jugador', 'Fecha', 'Hora', 'Puntos', 'Tipo'])
        else:
            log_df = pd.DataFrame(columns=['Jugador', 'Fecha', 'Hora', 'Puntos', 'Tipo'])

        fecha_hora = datetime.datetime.now()
        fecha = fecha_hora.date()
        hora = fecha_hora.time()

        log_entries = []

        for player, tipos in types.items():
            for tipo, puntos in tipos.items():
                log_entry = {
                    'Jugador': player,
                    'Fecha': fecha,
                    'Hora': hora,
                    'Puntos': puntos,
                    'Tipo': tipo
                }
                log_entries.append(log_entry)

        new_entries_df = pd.DataFrame(log_entries)
        log_df = pd.concat([log_df, new_entries_df], ignore_index=True)

        save_df(log_df, LOG_DIRECTORY)
        logger_info.info(f"Log creado correctamente en {LOG_DIRECTORY}")
    except Exception as e:
        logger_debug.error(f"Error creando log: {e}")


def log_statistic():
    try:
        if os.path.exists(LOG_DIRECTORY):
            log_df = pd.read_csv(LOG_DIRECTORY)
            logger_info.info("Log cargado correctamente para estadísticas.")
        else:
            log_df = pd.DataFrame(columns=['Jugador', 'Fecha', 'Hora', 'Puntos', 'Tipo'])
            logger_debug.warning("Archivo de log no encontrado. Se creó un DataFrame vacío.")

        group_log_df = log_df.groupby('Jugador')
        logger_info.info("Estadísticas agrupadas por jugador.")

    except Exception as e:
        logger_debug.error("Error creando estadística: {e}")


def backup_current_state():
    try:
        if os.path.exists(RESULTS_DIRECTORY):
            df = pd.read_csv(RESULTS_DIRECTORY)
            history_stack.append(df.copy())
            logger_info.info("Estado actual respaldado correctamente en la pila.")
        else:
            logger_debug.warning("No se encontró un estado previo para respaldar.")
    except Exception as e:
        logger_debug.error(f'Error al respaldar el estado actual: {e}')


def restore_previous_state():
    try:
        if history_stack:
            previous_df = history_stack.pop()
            save_df(previous_df, RESULTS_DIRECTORY)
            logger_info.info("Estado anterior restaurado correctamente desde la pila.")
            return jsonify({'status': 'success', 'message': 'Estado anterior restaurado exitosamente.'}), 200
        else:
            logger_info.warning("No hay mas estados disponibles.")
            return jsonify({'status': 'error', 'message': 'No hay más estados para deshacer.'}), 404
    except Exception as e:
        logger_debug.debug(f'Error al restaurar el estado anterior {e}')
        return jsonify({'status': 'error', 'message': 'Error al restaurar el estado anterior.'}), 500

def limpiar_texto(texto):
    """
    Removes unwanted Unicode characters from the input text.
    """
    # Define unwanted characters (e.g., Zero Width Space, Byte Order Mark, Word Joiner)
    unwanted_chars = ['\u200B', '\uFEFF', '\u2060']
    for char in unwanted_chars:
        texto = texto.replace(char, '')
    return texto

def procesar_resultados(texto):
    try:
        texto = limpiar_texto(texto)
        data = {}
        jugador_actual = None
        lineas = texto.splitlines()

        for idx, linea in enumerate(lineas, start=1):
                original_line = linea  # Save the original line for logging
                linea = linea.strip().upper()
                logger_debug.debug(f"Processing line {idx}: '{original_line}' -> '{linea}'")

                if re.match(r'^-', linea):
                    if jugador_actual:
                        # Remove the first hyphen and any leading characters
                        jugada = linea.lstrip('-').strip()
                        if jugada:
                            data[jugador_actual].append(jugada)
                            logger_debug.debug(f"Added jugada '{jugada}' to player '{jugador_actual}'")
                    else:
                        logger_debug.warning(f"Jugada found without a current player at line {idx}: '{linea}'")
                elif linea:
                    jugador_actual = linea
                    data[jugador_actual] = []
                    logger_debug.debug(f"Identified new player: '{jugador_actual}'")
                else:
                    logger_debug.warning(f"Empty or unexpected format at line {idx}: '{linea}'")

        # Si no se encontraron jugadores, retorna un DataFrame vacío
        if not data:
            logger_info.warning("No se encontraron jugadores válidos en los datos.")
            return pd.DataFrame(columns=['Jugador', 'Puntos']), [], [], []

        jugadores = list(data.keys())
        puntos = []
        tipos = {}

        for jugador in jugadores:
            p, t = calcular_puntos(data[jugador], jugador)
            for punto, tipo in zip(p, t):
                if jugador not in tipos:
                    tipos[jugador] = {}
                if tipo in tipos[jugador]:
                    tipos[jugador][tipo] += punto
                else:
                    tipos[jugador][tipo] = punto

            puntos.append(sum(p))

        df = pd.DataFrame({
            'Jugador': jugadores,
            'Puntos': puntos
        })

        logger_info.info(f"Resultados procesados: {df}")
        logger_info.info(f"Data procesados: {data}")
        logger_info.info(f"Tipos procesados: {tipos}")
        return df, tipos

    except Exception as e:
        logger_debug.error(f"Error processing results: {e}")
        return pd.DataFrame(columns=['Jugador', 'Puntos']), [], [], []


@app.route('/undo_last_update', methods=['POST'])
def undo_last_update():
    return restore_previous_state()


@app.route('/download_latest_results', methods=['GET'])
def download_latest_results():
    try:
        data = load_df(RESULTS_DIRECTORY)
        if data.empty:
            logger_info.warning("No hay respuesta disponible para descargar.")
            return jsonify({'status':'error', 'message': 'No hay respuesta disponible para descargar.'}), 404
        
        final_data = data.groupby('Jugador', as_index=False)[['Puntos', 'Victorias']].sum()
        final_data = final_data.sort_values('Puntos', ascending=False)

        aux_file = f'Resultados.{datetime.date.today()}.xlsx'
        aux_file_path = os.path.join(os.path.join('data', 'results'), aux_file)
        final_data.to_excel(aux_file_path, index=False)

        # Aplicar estilo al xlsx
        wb = load_workbook(aux_file_path)
        ws = wb.active
        ws.title = "Resultados"

        # Aplicar estilos a la tabla
        # Crear una tabla de Excel con estilo predeterminado
        tab = Table(displayName="Resultados", ref=ws.dimensions)

        # Estilo de la tabla
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Aplicar estilos a los encabezados
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        # Ajustar las dimensiones de la columna 'Jugador'
        ws.column_dimensions['A'].width = 28  # 'Jugador' column
        ws.column_dimensions['B'].width = 15  # 'Puntos' column
        ws.column_dimensions['C'].width = 15  # 'Victorias' column

        # Centrar los valores en la columna 'Puntos'
        for cell in ws['B'] + ws['C']:
            cell.alignment = alignment

        wb.save(aux_file_path)
        logger_info.info(f"Archivo de resultados generado y listo para descargar: {aux_file_path}")

        return send_file(aux_file_path, as_attachment=True, download_name=aux_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger_debug.error(f'Error en download_latest_results: {e}')
        return jsonify({'status': 'error', 'message': 'Error al generar el archivo para descargar.'}), 500

    
@app.route('/get_latest_results', methods=['GET'])
def get_latest_results():
    try:
        data = load_df(RESULTS_DIRECTORY)

        if data.empty:
            logger_info.warning("Error: No hay datos disponibles en el archivo.")
            return jsonify({'status': 'error', 'message': 'No hay resultados disponibles.'}), 404

        # Agrupación y ordenación de datos
        final_data = data.groupby('Jugador', as_index=False)[['Puntos', 'Victorias']].sum()
        final_data = final_data.sort_values(by='Puntos', ascending=False)

        logger_info.info("Resultados obtenidos y procesados correctamente.")
        response = make_response(jsonify({'status': 'success', 'data': final_data.to_dict(orient='records')}), 200)
        response.headers['Cache-Control'] = 'no-store'

        return response

    except Exception as e:
        logger_debug.error(f"Error in get_latest_results: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred.'}), 500


@app.route('/update_results', methods=['POST'])
def update_results():
    try:
        backup_current_state()
        
        content = request.json.get('results', '')
        if not content:
            logger_info.warning("No results provided in the request.")
            return jsonify({'status': 'error', 'message': 'No results provided.'}), 400

        new_data, types = procesar_resultados(content)

        new_data['Puntos'] = new_data['Puntos'].astype(int)
        data = load_df(RESULTS_DIRECTORY)

        final_data = pd.merge(data, new_data, on='Jugador', how='outer', suffixes=('_old', '_new'))
        final_data['Puntos'] = final_data['Puntos_old'].fillna(0) + final_data['Puntos_new'].fillna(0)
        final_data['Victorias'] = final_data['Victorias'].fillna(0)
        final_data = final_data[['Jugador', 'Puntos', 'Victorias']]

        save_df(final_data, RESULTS_DIRECTORY)
        create_log(types)

        logger_info.info("Resultados actualizados correctamente.")
        return jsonify({'status': 'success', 'message': 'Se actualizó correctamente', 'data': final_data.to_dict(orient='records')}), 200
    except Exception as e:
        logger_debug.error(f"Error in update_results: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred.'}), 500
    

@app.route('/update_table', methods=['POST'])
def update_table():
    try:
        backup_current_state()
        data = request.json.get('data', [])
        if not data:
            return jsonify({'status': 'error', 'message': 'No data provided.'}), 400
        
        df = pd.DataFrame(data)
        
        if 'Jugador' not in df.columns or 'Puntos' not in df.columns or 'Victorias' not in df.columns:
            return jsonify({'status': 'error', 'message': 'Invalid data format.'}), 400
        
        save_df(df, RESULTS_DIRECTORY)
        return jsonify({'status': 'success', 'message': 'Data updated successfully.'}), 200
    except Exception as e:
        logger_debug.error(f"Error in update_table: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred.'}), 500


if __name__ == '__main__':
    app.run(debug=True)